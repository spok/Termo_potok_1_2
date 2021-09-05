import numpy as np
from openpyxl import load_workbook
from datetime import datetime, timedelta
from canal import MyCanal
import math
import json

class MyData:
    def __init__(self):
        # Исходные данные из таблицы в виде списка
        self.data = np.array([])  # измерения по всем каналам
        self.data_dev = np.array([])  # отклонения от среднего значения
        self.date_time = []  # дата и время измерений
        self.date_time_move = []
        self.date_time_new = []
        # наименование каналов
        self.data_label = []
        self.data_count = []
        # статистические характеристики каналов
        self.data_min = []
        self.data_max = []
        self.data_mid = []
        # максимальные и минимальные значения
        self.max_value = 0.0
        self.min_value = 0.0
        # границы рабочего диапазона
        self.left_border = 0
        self.right_border = 0
        #
        self.word_file = ''
        self.time_step = 0  # значение текущего смещения по времени

        # распределение каналов
        self.canals = [MyCanal(i) for i in range(0, 3)]
        self.current_canals = 0

    def load_excel(self, fname: str) -> None:
        """Загрузка данных из файла Excel, fname - полный путь к файлу"""
        # ссылка на лист в таблице
        wb = load_workbook(fname)
        sheet = wb.worksheets[0]
        # количество заполненных строк
        max_row = sheet.max_row
        # заполнение наименований каналов
        self.data_label = []
        for j in range(0, 10):
            temp_str_name = str(sheet.cell(row=2, column=6 + j).value).split(',')
            self.data_label.append(temp_str_name[0])
        # запись времени измерений с преобразованеим в datetime
        for i in range(3, max_row + 1):
            first_date = datetime.strptime(str(sheet.cell(row=i, column=3).value), '%Y-%m-%d %H:%M:%S')
            self.date_time.append(first_date)
        self.date_time_move = self.date_time.copy()
        # запись 10 каналов датчиков в список
        data_spisok = []
        for j in range(0, 10):
            cur_data = []
            for i in range(3, max_row + 1):
                cur_cell_data = sheet.cell(row=i, column=6 + j).value
                if i == 3 or i == 4 or i == 5:
                    if cur_cell_data == 0:
                        cur_cell_data = 0.001
                cur_data.append(cur_cell_data)
            data_spisok.append(cur_data)

        # преобразование в массив numpy
        self.data = np.array(data_spisok)

        # настройка рабочего диапазона
        l = len(self.date_time)
        self.left_border = 0
        self.right_border = l

        self.calc_midle()

    def load_potokfile(self, fname: str, series: int) -> None:
        """Загрузка исходного файла потока, fname - полный путь к файлу"""
        f = open(fname)
        i = 1
        parse = []
        parse_text = []
        # чтение текстовых строк из файла
        for line in f:
            parse_text.append(line)
        # используется последняя строка с данными
        line = parse_text[-1]
        left = 0
        # парсинг строки с данными с сохранением структуры данных
        while left < len(line) - 1:
            parse_text = []
            # чтение номера серии
            left = left
            right = left + 19
            parse_text.append(float(line[left:right].lstrip()))
            # чтение даты измерения
            left = right
            right = left + 18
            parse_text.append(line[left:right])
            right = right + 136
            # чтение каналов
            for j in range(0, 10):
                left = right
                right = left + 18
                parse_text.append(float(line[left:right].lstrip()))
            # чтение типа каналов
            right = right + 194
            for j in range(0, 10):
                left = right
                right = left + 4
                parse_text.append(int(line[left:right].lstrip()))
            left = right
            # добавление записи по всем датчикам за один момент времени
            parse.append(parse_text)

        # количество измерений
        max_row = len(parse)
        # заполнение типов каналов
        self.data_label = []
        nq = 0
        nt = 0
        for j in range(12, 22):
            if parse[0][j] == 2:
                nq += 1
                self.data_label.append('q' + str(nq))
            if parse[0][j] == 1:
                nt += 1
                self.data_label.append('t' + str(nt))

        # чтение массива с датами измерений
        self.date_time = []
        find_series = []  # список списка с началом и концом каждой серии
        cur_series = 0.0
        f = []
        for i, line in enumerate(parse):
            # запись времени измерений с преобразованеим в datetime
            first_date = datetime.strptime(line[1], '%d.%m.%Y%H:%M:%S')
            self.date_time.append(first_date)

            # определение границ каждой серии
            cur = line[0]
            # на первом элементе списка
            if cur_series == 0.0:
                cur_series = cur
                f_left = i
            # при изменении номера серии
            if (cur_series != cur) or (i == (len(parse) - 1)):
                cur_series = cur
                f.append(f_left)
                if i == (len(parse) - 1):
                    f.append(i + 1)
                else:
                    f.append(i)
                find_series.append(f)
                f = []
                f_left = i

        # запись 10 каналов датчиков в список
        data_spisok = []
        for j in range(2, 12):
            cur_data = []
            for line in parse:
                if j > 1 and j < 5:
                    # защита от нулевых показаний датчиков потока
                    if line[j] == 0:
                        line[j] = 0.001
                cur_data.append(line[j])
            data_spisok.append(cur_data)

        # преобразование в массив numpy
        filter_data = np.array(data_spisok)
        # проверка номера серии на наличие в считанных данных, иначе использует 1 серию
        if series > len(find_series):
            series = 1
        # выборка данных для нужной серии измерений
        self.data = filter_data[:, find_series[series - 1][0]: find_series[series - 1][1]].copy()
        self.date_time = self.date_time[find_series[series - 1][0]: find_series[series - 1][1]]
        # настройка рабочего диапазона
        self.left_border = 0
        self.right_border = len(self.date_time)

        self.calc_midle()

    def get_border(self, name: str) -> int:
        """Возвращает значения границ рабочего диапазона, name - сторона диапазона left, right"""
        l = len(self.date_time)
        if name == 'left':
            if self.left_border == 0 and l > 0:
                self.left_border = 0
            return self.left_border
        else:
            if self.right_border == 0 and l > 0:
                self.right_border = l
            return self.right_border

    def get_count_element(self):
        """Возвращает общее количество элементов в списке с данными"""
        return len(self.date_time)

    def get_list_canal(self):
        """Возвращает список с типами каналов измерений"""
        return self.data_label

    def set_date_time(self, step: int = 0):
        """Возвращение списка с датами измерений со смещением на целое число дней,
        step - смещение в днях, целое"""
        # определяем смещение времени от предыдущего значения
        a = timedelta(days=(step - self.time_step))
        for i in range(0, len(self.date_time)):
            self.date_time[i] += a
        self.time_step = step

    def get_canal_data(self):
        """Возвращает список с данными каналов измерений"""
        return self.data

    def get_midle_canal(self, canal: int) -> float:
        """Возвращает среднее значение канала"""
        return self.data_mid[canal]

    def get_mid_canal(self, col: int, typ: str) -> float:
        """Возврат значения среднего или экстремумов"""
        if typ == 'mid':
            return self.data_mid[col]
        elif typ == 'min':
            return self.data_min[col]
        else:
            return self.data_max[col]

    def calc_midle(self) -> None:
        """Расчет средних значений и экстремумов данных"""
        # расчет параметров каждого канала
        self.data_max = []
        self.data_min = []
        self.data_mid = []
        for j in range(0, 10):
            # добавление элемента в список для дальнейшего расчета
            self.data_max.append(np.amax(self.data[j, self.left_border:self.right_border]))
            self.data_min.append(np.amin(self.data[j, self.left_border:self.right_border]))
            self.data_mid.append(np.mean(self.data[j, self.left_border:self.right_border]))
        # поиск минимума и максимума сразу всех каналов
        self.min_value = np.amin(np.array(self.data_min))
        self.max_value = np.amax(np.array(self.data_max))
        # расчет отклонений измерений канала от среднего значения канала
        self.data_dev = self.data.copy()
        for i in range(0, 10):
            for j in range(0, len(self.data_dev[i])):
                if self.data_mid[i] == 0:
                    self.data_dev[i, j] = 0
                else:
                    self.data_dev[i, j] = abs((1.0 - self.data_dev[i, j] / self.data_mid[i]) * 100)

    def get_border_y(self, rez: list) -> list:
        """Определение экстремумов массива измерений"""
        min = np.amin(rez)
        max = np.amax(rez)
        return [min, max]

    def edit_cell_data(self, col: int, row: int, val: float):
        """Изменение значения измерений c пересчетом"""
        self.data[col, row] = val
        self.calc_midle()

    def copy_data(self):
        """копирование данных измерений для каждой конструкции"""
        for con in self.canals:
            con.canal = np.empty((4, self.right_border - self.left_border))
            con.ro = np.empty((2, self.right_border - self.left_border))
            for j in range(0, 4):
                num = con.canal_number[j]
                con.canal[j] = self.data[num, self.left_border:self.right_border].copy()

    def calc_pogr(self, con):
        """Расчет погрешности измерений"""
        delta_t = con.canal[2] - con.canal[3]
        t_sr = np.mean(delta_t)
        delta_t = abs(delta_t - t_sr)
        sr_delta_t = np.mean(delta_t)
        q_sr = np.mean(con.canal[0])
        delta_q = abs(con.canal[0] - q_sr)
        delta_q_sr = np.mean(delta_q)
        con.dr = con.ro1 * math.sqrt(delta_q_sr ** 2 / q_sr ** 2 + sr_delta_t ** 2 / t_sr ** 2)

    def calc_r(self):
        """Расчет сопротивлений теплопередаче для всех конструкций с учетом всех коэффициентов"""
        self.copy_data()
        for con in self.canals:
            # Пересчет значений каналов каждой конструкции
            for j in range(0, 4):
                base = self.data_mid[con.canal_number[j]]
                delta = con.canal_koef[j][0] + con.canal_koef[j][1]
                con.canal_value[j] = base + delta
                con.canal[j] = con.canal[j] + delta
                # пересчет с учетом масштабного коэффициента
                koef = con.canal_koef[j][2]
                if koef > 1:
                    median = np.mean(con.canal[j])
                    for i in range(0, len(con.canal[j])):
                        con.canal[j][i] = median + (con.canal[j][i] - median) / koef
            # Расчет сопротивлений теплопередаче
            if con.canal_value[0] == 0:
                ro1 = 0
                ro2 = 0
            else:
                ro1 = abs((con.canal_value[2] - con.canal_value[3]) / con.canal_value[0])
                ro2 = abs((con.canal_value[1] - con.canal_value[3]) / con.canal_value[0] + 1 / 8.7)
            con.ro1 = ro1
            con.ro2 = ro2
            # Расчет массива сопротивлений теплопередаче
            con.ro[0] = abs((con.canal[2] - con.canal[3]) / con.canal[0])
            con.ro[1] = abs((con.canal[1] - con.canal[3]) / con.canal[0] + 1 / 8.7)
            self.calc_pogr(con)
        self.calc_midle()

    def save_base(self, fname: str) -> str:
        """Сохраненение обработанных данных"""
        if fname != '':
            save_data = {}
            save_data['sourse_label'] = self.data_label
            buf = [datetime.strftime(i, '%Y.%m.%d %H:%M:%S') for i in self.date_time]
            save_data['sourse_time'] = buf
            # сохранение исходных данных с измерениями
            save_data['sourse_data'] = self.data.tolist()
            # Сохранение настроек каналов и конструкций
            konstr_list = []
            for i, con in enumerate(self.canals):
                temp_constr = {}
                temp_constr['name'] = con.name
                temp_constr['canal_number'] = con.canal_number
                temp_constr['canal_koef'] = con.canal_koef
                konstr_list.append(temp_constr)
            save_data['property'] = konstr_list
            save_data['left_border'] = self.left_border
            save_data['right_border'] = self.right_border
            save_data['data_step'] = self.time_step
            try:
                with open(fname, 'w') as save_file:
                    json.dump(save_data, save_file)
                return 'OK'
            except:
                return 'Не удалось записать файл'

    def load_base(self, fname: str) -> str:
        """Загрузка сохраненных обработанных данных"""
        if fname != '':
            try:
                with open(fname, 'r') as load_file:
                    load_data = json.load(load_file)
                self.data_label = load_data['sourse_label']
                # Загрузка времени измерений
                buf = load_data['sourse_time']
                for i in buf:
                    self.date_time.append(datetime.strptime(i, '%Y.%m.%d %H:%M:%S'))
                # Загрузка исходных данных с измерениями
                self.data = np.array(load_data['sourse_data'])
                # Сохранение настроек каналов и конструкций
                konstr_list = load_data['property']
                for i, con in enumerate(konstr_list):
                    self.canals[i].name = con['name']
                    self.canals[i].canal_number = con['canal_number']
                    self.canals[i].canal_koef = con['canal_koef']

                self.left_border = load_data['left_border']
                self.right_border = load_data['right_border']
                self.time_step = load_data['data_step']
                self.calc_midle()
                return 'OK'
            except:
                return 'Не удалось открыть файл'
